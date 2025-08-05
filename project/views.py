from django.shortcuts import render,redirect
from .decorators import project_required
from accounts.models import *
from django.contrib.auth import authenticate,logout
from django.http import HttpResponseBadRequest,JsonResponse,HttpResponse
import pandas as pd
import openpyxl
from . utilities import *
from decimal import Decimal
from django.core.paginator import Paginator
from django.db.models import Sum
from openpyxl.styles import Font,PatternFill,Alignment,colors,Border,Side,NamedStyle
from openpyxl.drawing.image import Image
from django.conf import settings
from django.contrib.auth.hashers import make_password
from django.utils.timezone import now
from django.contrib import messages



@project_required
def dashboard(request):
    try:
        # companies=''
        # consultants=''
        # owners=''
        currentuser=request.user
        projectobjs=Project.objects.values_list('user', flat=True)
        print(projectobjs)        
        if currentuser.id in projectobjs:
            projectobj=Project.objects.get(user=currentuser)
            bills=Bill.objects.filter(project=projectobj)
            activebill=Bill.objects.filter(project=projectobj,submitted=False).count()
            wo_boq_uploaded=WOBOQItem.objects.filter(project=projectobj).exists()
            if wo_boq_uploaded:
                boq_approve_count=Pending_Item.objects.filter(project=projectobj).count()
            else:
                boq_approve_count=0
            context={
                'registered' :True,
                'workorder_uploaded':wo_boq_uploaded,
                'projectobj' : projectobj, 
                'bills':bills,
                'billcount':bills.count(),
                'activebill':activebill,
                'boq_approve_count':boq_approve_count           
            }            
        else:            
            context={
                'registered' :False,                
            }
        return render(request,'project/dashboard.html',context)
    except Exception as e:
        context={
                'registered' :'Error',            
            }
        messages.error(request, "Something went wrong : "+str(e))
        return render(request,'project/dashboard.html',context) 
    
@project_required
def ajax_get_owners(request):
    query = request.GET.get('q', '')
    if query and len(query) >= 4:
        # Case-insensitive search for matching usernames
        owners = Owner.objects.filter(owner_name__icontains=query).values('id', 'owner_name')#[:10]  # Limit results
        return JsonResponse(list(owners), safe=False)
    return JsonResponse([], safe=False) 

from django.shortcuts import get_object_or_404
@project_required
def ajax_get_consultants(request):
    owner_id = request.GET.get('owner_id')  # Retrieve the owner_id from the GET request
    if owner_id:
        # Query consultants associated with the owner (adjust according to your model)
        owner=Owner.objects.get(id=owner_id)
        consultants=[]
        if owner:
            consultants = Consultant.objects.filter(owner=owner)
        # Prepare data to send back as JSON
        data = [{'id': consultant.id, 'consultant_name': consultant.consultant_name} for consultant in consultants]
        print('consultants found>>>>>>>')
        print(data)
        return JsonResponse(data, safe=False)
    return JsonResponse({'error': 'No owner_id provided'}, status=400)

@project_required
def ajax_get_companies(request):
    query = request.GET.get('q', '')
    if query and len(query) >= 4:
        # Case-insensitive search for matching usernames
        companies = Company.objects.filter(company_name__icontains=query).values('id', 'company_name')#[:10]  # Limit results
        print('companies found>>>>>>>')
        print(companies)
        return JsonResponse(list(companies), safe=False)
    return JsonResponse([], safe=False)


@project_required
def add_consultant(request):
    current_user=request.user    
    msg=''
    try:
        if request.method == 'POST':            
            name=request.POST.get('consultantcompanyname')
            emailid=request.POST.get('consultantemailid')
            username,password=get_username_password(name)
            print('username =',username)
            print('password =',password)
            address=request.POST.get('consultantaddress')
            country_code=request.POST.get('consultantcountrycode')
            phone=request.POST.get('consultantphone')
            owner_id=request.POST.get('ownerid')
            print('OWNER ID=',owner_id)
            owner=Owner.objects.get(id=owner_id)
            
            if User.objects.filter(username=username).exists():
                return JsonResponse({'error': 'Already exists.'}, status=400)               
            else:
                with transaction.atomic():
                    epassword=make_password(password)
                    user=User.objects.create(first_name=name,username=username,password=epassword,is_staff=False,email=emailid,user_type='consultant',added_by=current_user.username,is_active=False)
                    print('new user=',user)
                    newconsultant=Consultant.objects.create(
                    user=user,
                    consultant_name=name,
                    consultant_address=address,
                    consultant_email=emailid,
                    country_code=country_code,
                    consultant_phone=phone,
                    owner=owner
                                    
                    )  
                    
                    recipient_list = [emailid] 
                    context = {
                        'consultantusername': username,
                        'consultantname': name,
                        'password': password,
                        'contractor': current_user.first_name,
                    }
                    transaction.on_commit(lambda: send_user_email.delay(context,'consultant',recipient_list))

                        
                    return JsonResponse({'message': 'Consultant added successfully!','id':newconsultant.id,'name':newconsultant.consultant_name})
        return JsonResponse({'error': 'Invalid request method.'}, status=400) 
    except Exception as e:
        messages.error(request, "Error : "+str(e))
        print(e)                
    return JsonResponse({'error': 'Invalid request method.'}, status=400)    

@project_required
def add_company(request):
    current_user=request.user    
    msg=''
    try:
        if request.method == 'POST':            
            name=request.POST.get('contractorcompanyname')
            emailid=request.POST.get('companyemailid')
            username,password=get_username_password(name)
            print('username =',username)
            print('password =',password)
            address=request.POST.get('companyaddress')
            country_code=request.POST.get('companycountrycode')
            phone=request.POST.get('companyphone')
            
            if User.objects.filter(username=username).exists():
                return JsonResponse({'error': 'Already exists.'}, status=400)               
            else:
                with transaction.atomic():
                    epassword=make_password(password)
                    user=User.objects.create(first_name=name,username=username,password=epassword,is_staff=False,email=emailid,user_type='company',added_by=current_user.username,is_active=False)
                    print('new user=',user)
                    newcompany=Company.objects.create(
                    user=user,
                    company_name=name,
                    company_address=address,
                    company_email=emailid,
                    country_code=country_code,
                    company_phone=phone,                                    
                    )  
                    
                    recipient_list = [emailid] 
                    context = {
                        'companyusername': username,
                        'companyname': name,
                        'password': password,
                        'projectname': current_user.first_name,
                    }
                    transaction.on_commit(lambda: send_user_email.delay(context,'company',recipient_list))

                        
                    return JsonResponse({'message': 'Company added successfully!','id':newcompany.id,'name':newcompany.company_name})
        return JsonResponse({'error': 'Invalid request method.'}, status=400) 
    except Exception as e:
        messages.error(request, "Error : "+str(e))
        print(e)                
    return JsonResponse({'error': 'Invalid request method.'}, status=400)        


@project_required
def add_owner(request):
    current_user=request.user    
    msg=''
    try:
        if request.method == 'POST':            
            name=request.POST.get('ownercompanyname')
            emailid=request.POST.get('owneremailid')
            username,password=get_username_password(name)
            print('username =',username)
            print('password =',password)
            address=request.POST.get('owneraddress')
            country_code=request.POST.get('ownercountrycode')
            phone=request.POST.get('ownerphone')
            
            if User.objects.filter(username=username).exists():
                return JsonResponse({'error': 'Already exists.'}, status=400)               
            else:
                with transaction.atomic():
                    epassword=make_password(password)
                    user=User.objects.create(first_name=name,username=username,password=epassword,is_staff=False,email=emailid,user_type='owner',added_by=current_user.username,is_active=False)
                    print('new user=',user)
                    newowner=Owner.objects.create(
                    user=user,
                    owner_name=name,
                    owner_address=address,
                    owner_email=emailid,
                    country_code=country_code,
                    owner_phone=phone,
                                    
                    )  
                    
                    recipient_list = [emailid] 
                    context = {
                        'ownerusername': username,
                        'ownername': name,
                        'password': password,
                        'contractor': current_user.first_name,
                    }
                    transaction.on_commit(lambda: send_user_email.delay(context,'owner',recipient_list))

                    print('owner added =',newowner)           
                    return JsonResponse({'message': 'Owner added successfully!','id':newowner.id,'name':newowner.owner_name})
        return JsonResponse({'error': 'Invalid request method.'}, status=400) 
    except Exception as e:
        messages.error(request, "Error : "+str(e))
        print(e)                
    return JsonResponse({'error': 'Invalid request method.'}, status=400)             

@project_required
def manage_project(request):
    if request.method == 'POST':
        try:
            project_name=request.POST.get('name') 
            
            company_id=request.POST.get('companyId')
            consultant_id=request.POST.get('consultant')
            owner_id=request.POST.get('ownerId')
            location=request.POST.get('location')
            other_details=request.POST.get('otherdetails')
            company=Company.objects.get(id=company_id)
            owner=Owner.objects.get(id=owner_id)
            consultant=Consultant.objects.get(id=consultant_id)
                        
            newproject=Project.objects.create(
                user=request.user,
                project_name=project_name,
                consultant=consultant,
                owner=owner,
                company=company,
                location=location,
                other_details=other_details                
            ) 
            # print(newproject)  
            # print("newproject details added.....") 
        except Exception as e:
            messages.error(request, "Error : "+str(e))                    
    return redirect('project:Dashboard')

@project_required
def edit_project(request):
    current_user=request.user
    projects=Project.objects.filter(user=current_user)
    if projects:
        project=projects[0]
    else:
        messages.error(request, "Error :No Project Found ") 
        return redirect('project:Dashboard')
    if request.method == 'POST':
        try:
            project_name=request.POST.get('name') 
            
            company_id=request.POST.get('companyId')
            consultant_id=request.POST.get('consultant')
            owner_id=request.POST.get('ownerId')
            location=request.POST.get('location')
            other_details=request.POST.get('otherdetails')
            company=Company.objects.get(id=company_id)
            owner=Owner.objects.get(id=owner_id)
            consultant=Consultant.objects.get(id=consultant_id)                     
            
            
            project.project_name=project_name
            project.consultant=consultant
            project.owner=owner
            project.company=company
            project.location=location
            project.other_details=other_details  
            project.save()              
            
            # print(newproject)  
            # print("newproject details added.....") 
        except Exception as e:
            messages.error(request, "Error : "+str(e))          
            
        return redirect('project:Dashboard')
    return render(request,'project/edit_project.html',{'project':project})

@project_required
def change_password(request):
    msg=''
    if request.method == 'POST':
        user_name=request.POST.get('username')
        existingpw=request.POST.get('oldpassword')
        newpw=request.POST.get('newpassword')
        newpw2=request.POST.get('confirmpassword')
        # print("Credentials given by user>>>>>",user_name,existingpw)
        if newpw == newpw2 :
            if not User.objects.filter(username=user_name).exists():
                messages.error(request, 'Failed, Username does not exist !') 
                                
            elif request.user.username != user_name:
                messages.error(request, 'Failed, Username mismatch !') 
                                
            else:
                try:
                    user = authenticate(request, username=user_name, password=existingpw)
                    # print("Authenticated>>>>>>",user.username)
                    if user is not None  and user == request.user:                        
                        user.set_password(newpw)
                        user.save()
                        # print('Successfully changed to new pwd>>>>',newpw)
                        return redirect('accounts:UserLogin')
                    else:
                        msg='Failed, try again with correct login credentials !'                        
                except Exception as e:
                    messages.error(request, "Error : "+str(e))
                    
        else:
            messages.error(request, 'Failed,Password mismatch....') 
                  
    return render(request,'project/change_password.html')


@project_required
def userlogout(request):
    try:
        logout(request)
        return redirect('accounts:UserLogin')
    except:
        return redirect('accounts:UserLogin')


    
def get_heading_row(xlfile):
    temp_df = pd.read_excel(xlfile, nrows=20, header=None)  
    for i, row in temp_df.iterrows():
        if 'AMOUNT' in row.values or 'DESCRIPTION' in row.values or 'Item Description' in row.values or 'Description' in row.values or 'Quantity' in row.values: 
            return i
    return 0

@project_required
def ajax_get_headings(request):
    headings=''
    if request.method == 'POST' and request.FILES.get('xlfile'):
        workorder_file = request.FILES['xlfile']
        headingrow=get_heading_row(workorder_file)
        df = pd.read_excel(workorder_file, skiprows=headingrow, header=0)
        headings=df.columns.tolist()
        # print(">>>Headings>>>>>>",headings)
    return JsonResponse({'headings':headings})



from celery.result import AsyncResult

@project_required
def get_upload_status(request):
    task_id = request.GET.get('task_id')
    result = AsyncResult(task_id)
    if result.status=='SUCCESS':
        if WOBOQItem.objects.filter(project__user=request.user).exists():
            status='SUCCESS'
        else:
            status='FAILURE'
        return JsonResponse({'status': status})
    else:
        return JsonResponse({'status': result.status})
    


    
            
@project_required
def upload_workorder_boq(request):
    msg=''
    boqpending=True
    context={
        'msg' : msg,
        'boqpending':boqpending
    }
    if request.method == 'POST' and request.FILES.get('wo_boq_xl'):
        try:
            work_order_no=request.POST.get('wo_no')
            Project.objects.filter(user=request.user).update(work_order_no=work_order_no)
            dateofentry =request.POST.get('wo_date')
            sno_heading =request.POST.get('serialno')
            desc_heading=request.POST.get('description')
            qty_heading=request.POST.get('quantity')
            unit_heading=request.POST.get('unit')
            rate_heading=request.POST.get('rate')
            amt_heading=request.POST.get('amount')
            workorder_file = request.FILES['wo_boq_xl']

            file_content = workorder_file.read() 
            Project.objects.filter(user=request.user).update(work_order_no=work_order_no)           
            userid=request.user.id
            arg_dict={
                'userid':userid,
                'file_content':file_content,
                'sno_heading':sno_heading,
                'desc_heading':desc_heading,
                'qty_heading':qty_heading,
                'unit_heading':unit_heading,
                'rate_heading':rate_heading,
                'amt_heading':amt_heading,
                'dateofentry':dateofentry

            }
            task=process_boq_excel.delay(arg_dict)
            print('celery task id=',task.id)
            return render(request,'project/upload.html',{'task_id':task.id})            
        except Exception as e:
            messages.error(request, "Error : "+str(e))
            
        return render(request,'project/error_page.html',{'error_message': str(e)})           
    
    


@project_required
def manage_bill_of_quantities(request,page):
    currentuser=request.user
    msg=''
    projs=Project.objects.filter(user=currentuser)
    if projs:
        project=projs[0]
    else:
        messages.error(request, "Error :No project found ")
        return redirect('project:Dashboard')
    if request.method == 'POST':
        try:
            subheading=None
            # if request.POST.get('newsubhead') and request.POST.get('newserialno'):
            #     subheading=str(request.POST.get('newserialno')).strip()+'&&&'+str(request.POST.get('newsubhead')).strip()
            # print('subheading=',subheading) 

            description=request.POST.get('itemdesc')  
            serialno=request.POST.get('serialno') 
            
            wo_date=request.POST.get('wo_date')
            quantity=request.POST.get('quantity')
            rate=request.POST.get('rate')
            quantity=Decimal(quantity)
            rate=Decimal(rate)
            amount =quantity * rate
            unit=request.POST.get('unit')
            # adding to pending table, for getting approval from consultant and owner
            itemobject = Pending_Item(
                    project=project,
                    serial_no=serialno,
                    description=description,
                    rate=rate ,
                    quantity=quantity ,
                    unit=unit ,
                    amount=amount,
                    subheading=subheading,
                    extra_item=True,
                    work_order_date=wo_date,
                    heading=False,
                    existing=0
                )
            itemobject.save() 
            project =itemobject.project
            consultant=project.consultant
            owner=project.owner
            company=project.company
            desc='Added Extra Item to BoQ'
            newitem=desc='Added Item :'+ itemobject.description + ', Rate ='+str(itemobject.rate)+', Quantity = '+str(itemobject.quantity)
            recipient_list = [consultant.consultant_email,owner.owner_email] 
            cc_list=[company.company_email]
            emailcontext = {
                'name': 'Hi',
                'description': desc,
                'newitem': newitem,
                'projectname': project.project_name,
                'companyname':company.company_name,
                
            }
            item_approval_email.delay(emailcontext,recipient_list,cc_list)
            messages.info(request,'Item sent for Consultant and Owner approval.' )
            
        except Exception as e:
            messages.error(request, "Error : "+str(e))
            
        
    page_obj=None
    grandtotal=0.00
    allitemobjs=WOBOQItem.objects.filter(project__user=currentuser)
    if allitemobjs:
        total=allitemobjs.aggregate(totalamount=Sum('amount'))
        if total['totalamount'] is not None:                     
            grandtotal=total['totalamount'] 
        page_obj=get_boq(project,page) 
                 
    else: 
        messages.info(request,'No items found in BoQ, Add items to BoQ')       
        
    context={
       
        'page_object':page_obj,
        'grandtotal':grandtotal
       
    }
    return render(request,'project/manage_boq.html',context)


@project_required
def delete_bill_of_quantities(request):
    currentuser=request.user    
    try:
        projs=Project.objects.filter(user=currentuser)
        if projs:
            project=projs[0]
            return render(request,'project/confirm_delete.html',{'project':project})
        else:
            messages.error(request, "Error : No project found.")
            
    except Exception as e:
        messages.error(request, "Error : "+str(e))
    return render(request,'project/confirm_delete.html',{'project':None})
    
@project_required
def confirm_delete_boq(request,id):
    try:
        with transaction.atomic():            
            Project.objects.filter(id=id).update(work_order_no=None)   
            Bill.objects.filter(project__id=id).delete()
            Pending_Item.objects.filter(project__id=id).delete()
            WOBOQItem.objects.filter(project__id=id).delete()
        return render(request,'project/confirm_delete.html',{'project':None})
    except Exception as e:
        messages.error(request, "Error : "+str(e))
    return render(request,'project/confirm_delete.html',{'project':None})

@project_required
def manage_boq_item(request,itemid):
    
    boqitem=None
    try:
        itembjs=WOBOQItem.objects.filter(id=itemid)
        if itembjs:
            boqitem=itembjs[0]
            existing_qty = boqitem.quantity
            existing_rate = boqitem.rate
            existing_desc=boqitem.description
            if request.method == 'POST':
                subheading=None
                # if request.POST.get('newsubhead') and request.POST.get('newserialno'):
                #     subheading=str(request.POST.get('newserialno')).strip()+'&&&'+str(request.POST.get('newsubhead')).strip()
                #     boqitem.subheading=subheading
                if request.POST.get('itemdesc'):
                    description=request.POST.get('itemdesc')  
                else:
                    description=boqitem.description
                    
                if request.POST.get('serialno'):
                    serial_no=request.POST.get('serialno') 
                else:
                    serial_no=boqitem.serial_no
                
                if request.POST.get('extraitem'):                    
                    if request.POST.get('extraitem') == 'Yes':
                        extra_item=True
                    else:
                        extra_item=False
                else:
                    extra_item=boqitem.extra_item

                if request.POST.get('wo_date'):
                    work_order_date=request.POST.get('wo_date')
                else:
                    work_order_date=boqitem.work_order_date

                if request.POST.get('quantity'):
                    quantity=request.POST.get('quantity')
                else:
                    quantity=boqitem.quantity

                if request.POST.get('rate'):
                    rate=request.POST.get('rate')
                else:
                    rate=boqitem.rate

                if request.POST.get('amount'):
                    amount=request.POST.get('amount')
                else:
                    amount=boqitem.amount
                quantity=Decimal(quantity)
                rate=Decimal(rate)
                

                if request.POST.get('unit'):
                    unit=request.POST.get('unit')
                else:
                    unit=boqitem.unit                
                if quantity != existing_qty or rate != existing_rate or existing_desc != description:
                    itemobject = Pending_Item(
                        project=boqitem.project,
                        serial_no=serial_no,
                        description=description,
                        rate=rate ,
                        quantity=quantity ,
                        unit=unit ,
                        amount=amount,
                        subheading=None,
                        extra_item=extra_item,
                        work_order_date=work_order_date,
                        heading=boqitem.heading,
                        existing=boqitem.id
                    )
                    itemobject.save() 
                    project =itemobject.project
                    consultant=project.consultant
                    owner=project.owner
                    company=project.company
                    desc='Modified Item :'+ boqitem.description + ', Rate ='+str(existing_rate)+', Quantity = '+str(existing_qty)
                    newitem=description + ', Rate ='+str(itemobject.rate)+', Quantity = '+str(itemobject.quantity)
                    recipient_list = [consultant.consultant_email,owner.owner_email] 
                    cc_list=[company.company_email]
                    emailcontext = {
                        'name': 'Hi',
                        'description': desc,
                        'newitem': newitem,
                        'projectname': project.project_name,
                        'companyname':company.company_name,
                        
                    }
                    item_approval_email.delay(emailcontext,recipient_list,cc_list)
                    messages.info(request,'Item update sent for consultant and owner approval.' )
                    
                else:  
                    
                    boqitem.serial_no=serial_no
                    boqitem.description=description                    
                    boqitem.unit=unit 
                    boqitem.amount=amount                    
                    boqitem.extra_item=extra_item
                    boqitem.work_order_date=work_order_date                    
                    boqitem.save()   
                     
                    messages.info(request,'Item updated .' )
                
            
        else:
            messages.info(request,'Item not found .' )
    except Exception as e:
        messages.error(request, "Error : "+str(e))
        
    context={
                'boqitem':boqitem,
                
            }
    return render(request,'project/boq_item.html',context)


@project_required
def view_boq_approvals(request):
    projectobjs=Project.objects.filter(user=request.user)
    if projectobjs:
        project=projectobjs[0]
    else:
        messages.info(request,"project not found !")
    pendingobjs=None
    try:        
        if Pending_Item.objects.filter(project=project).exists():
            pendingobjs=Pending_Item.objects.filter(project=project)
        
    except Exception as e:
        pendingobjs=None
        messages.error(request, "Error : "+str(e))
    return render(request,'project/boq_approvals.html',{'pendingobjs':pendingobjs})

@project_required
def remove_boq_update(request,id):
    pendingobjs=None
    try:        
        pendingobj=Pending_Item.objects.get(id=id)
        project=pendingobj.project
        pendingobj.delete()
        messages.info(request,"Update Removed !")
        if Pending_Item.objects.filter(project=project).exists():
            pendingobjs=Pending_Item.objects.filter(project=project)        
    except Exception as e:
        try:
            if Pending_Item.objects.filter(project=project).exists():
                pendingobjs=Pending_Item.objects.filter(project=project)
        except:
            pass
        messages.error(request, "Error : "+str(e))
    return render(request,'project/boq_approvals.html',{'pendingobjs':pendingobjs})

@project_required
def remove_boq_item(request,itemid):
    msg=''
    try:
        itembjs=WOBOQItem.objects.filter(id=itemid)        
        if itembjs:

            boqitem=itembjs[0]
            if Measurement.objects.filter(woboqitem=boqitem).exists():
                messages.info(request,'Measurements exists for this item.' )
            else:
                Pending_Item.objects.filter(existing=boqitem.id).delete()
                boqitem.delete()
                messages.info(request,'Item removed .' )
        else:
            messages.info(request,'Item not found .' )            
    except Exception as e:
        messages.error(request, "Error : "+str(e))
        
    context={
        'boqitem':None,
        
    }
    return render(request,'project/boq_item.html',context)

#bills

def get_next_bill_number(project):
    bills=Bill.objects.filter(project=project)
    if bills.count() == 0:
        return 1
    else:
        return bills.count()+1
    
def ActiveBillExists(project):
    if Bill.objects.filter(project=project,submitted=False).exists():
        return True
    else:
        return False

@project_required
def get_billupload_status(request):
    task_id=request.GET.get('task_id')
    result = AsyncResult(task_id)
    print(">>>>>>>>>>>>>>>>>>>>>STATUS=",result.status)
    if result.status=='SUCCESS':
        if Measurement.objects.filter(bill__project__user=request.user).exists():
            status='SUCCESS'
        else:
            status='FAILURE'
        return JsonResponse({'status': status})
    else:
        return JsonResponse({'status': result.status})
    
@project_required
def create_bill(request): 
    try:   
        currentuser=request.user
        projectobjs=Project.objects.filter(user=currentuser)
        if projectobjs:
            project=projectobjs[0]
            if not WOBOQItem.objects.filter(project=project).exists():
                messages.info(request,'No work-order BOQ found. upload and try again.')            
                return redirect('project:Dashboard')
            if is_not_first_bill(project):
                return redirect('project:CreateNewBill')
            else:
                if request.method == 'POST':
                    if request.POST.get('projectbill') == 'subsequent':
                        currentbillno=request.POST.get('billNumber')
                        sno_heading =request.POST.get('serialno')
                        qty_heading=request.POST.get('quantity')
                        desc_heading=request.POST.get('description')                    
                        previousbillxl = request.FILES['previousbillxl']

                        file_content = previousbillxl.read() 
                        
                        arg_dict={
                            'projectid':project.id,
                            'file_content':file_content,
                            'sno_heading':sno_heading,
                            'desc_heading':desc_heading,
                            'qty_heading':qty_heading,
                            'currentbillno':currentbillno  

                        }
                        task=process_bill_excel.delay(arg_dict)
                        print('celery task id=',task.id)
                        return render(request,'project/upload_bill.html',{'task_id':task.id})  
                    else:
                        return redirect('project:CreateNewBill')                 
                else:
                    return render(request,'project/bill_position.html')
    except Exception as e:
        messages.error(request, "Error : "+str(e))
        return redirect('project:Dashboard')



@project_required
def create_new_bill(request):
    try:    
        currentuser=request.user
        projectobjs=Project.objects.filter(user=currentuser)
        project=projectobjs[0]  
        
        if request.method == 'POST':
            if ActiveBillExists(project):
                messages.info(request,"one bill is already active. Create new bill only after submitting existing bills.")
            else:
                bill_number=get_next_bill_number(project)
                bill_name=request.POST.get("name")
                bill_date=request.POST.get("bill_date")
                other_details=request.POST.get("otherdetails")
                try:
                    newbill=Bill.objects.create(
                        project=project,
                        bill_name=bill_name,
                        bill_number=bill_number,
                        bill_date=bill_date,
                        other_details=other_details
                    )
                    print('new bill created',newbill)
                    return redirect('project:CurrentBill',1)
                except Exception as e:
                    messages.error(request, "Error : "+str(e))                    
        
    except Exception as e:
        messages.error(request,'Something went wrong, try again !')
    return render(request,'project/create_bill.html')

'''
first- check if project details exists
second-check if unsubmitted bill exists, if yes
proceed and allow to modify items and measurements f that bill.
'''
@project_required
def current_bill(request,page):    
    try:
        msg=''
        current_bill=None
        currentuser=request.user
        projectobjs=Project.objects.filter(user=currentuser)
        if projectobjs:
            project=projectobjs[0]
        else:
            return redirect('project:Dashboard')

        if ActiveBillExists(project):
            currentbill_objs=Bill.objects.filter(project=project,submitted=False)
            current_bill=currentbill_objs[0]    
        else:
            messages.info(request," No bill is active. Create new bill to proceed.")
        page_obj=None
        
        if WOBOQItem.objects.filter(project__user=currentuser).exists():
            page_obj=get_boq(project,page) 
                        
        else:        
            messages.info(request," No items in boq. ")
        context={
            
            'current_bill':current_bill,
            
            'page_object':page_obj,
        
        }
        
        
        return render(request,'project/current_bill.html',context)
    except Exception as e:
        messages.error(request, "Error : "+str(e))
        
    return render(request,'project/current_bill.html')


@project_required
def share_to_consultant(request,billid):
    try:
        bill=Bill.objects.get(id=billid)
        consultant=bill.project.consultant
        consultant_email=consultant.consultant_email  
        company=bill.project.company 
        owner=bill.project.owner 
        current_time=now()  
        current_time_ist=get_time_in_ist(current_time)    
        bill.shared_to_consultant=current_time
        bill.save()
        recipient_list = [consultant_email,owner.owner_email,company.company_email] 
        context = {
            'name': 'Hi',
            'sharedtoname': consultant.consultant_name,
            'sharedbyname': company.company_name,
            'projectname': bill.project.project_name,
            'billname':bill.bill_name,
            'time':current_time_ist
        }
        send_email_notification.delay(context,recipient_list)
        
        
    except Exception as e:
        messages.error(request, "Error : "+str(e))
    return redirect('project:CurrentBill',1)


  
       
@project_required
def view_owner_comments(request,measureid):
    try:
        measurement=Measurement.objects.get(id=measureid)
        comments=Owner_Comments.objects.filter(measurement=measurement)
        context={
            'measurement':measurement,
            'comments':comments
        }
        try:
            Owner_Comments.objects.filter(measurement=measurement,read_by_project=False).update(read_by_project=True)
            
        except Exception as e:
            messages.error(request, "Error : "+str(e))
            
        return render(request,'project/owner_comments.html',context)
    except Exception as e:
        messages.error(request, "Error : "+str(e))
    return render(request,'project/owner_comments.html')    
       

@project_required
def search_boq_item(request,page): 
    try:   
        msg=''
        current_bill=None
        currentuser=request.user
        projectobjs=Project.objects.filter(user=currentuser)
        if projectobjs:
            project=projectobjs[0]
        else:
            messages.info(request," No project found. Add project and try again.")
        if ActiveBillExists(project):
            currentbill_objs=Bill.objects.filter(project=project,submitted=False)
            current_bill=currentbill_objs[0]    
        else:
            messages.info(request," No active bill. ")
        page_obj=None
        
        if WOBOQItem.objects.filter(project__user=currentuser).exists():
            searchterm=request.POST.get('searchitem')
            allrecs=WOBOQItem.objects.filter(project=project,description__icontains=searchterm).order_by('id')
            if allrecs:
                allrecs_paginator=Paginator(allrecs,len(allrecs)) 
                page_obj=allrecs_paginator.get_page(page)
                    
        else:        
            messages.info(request," No items in boq. ")
        context={
            
            'current_bill':current_bill,        
            'page_object':page_obj,       
        }
        
        return render(request,'project/current_bill.html',context)
    except Exception as e:
        messages.error(request, "Error : "+str(e))
    return render(request,'project/current_bill.html') 





@project_required
def manage_measurements(request,itemid,page):
    try:
        
        current_quantity = 0.00   
        page_object=None
        last_measure_date=None
        project=Project.objects.get(user=request.user)
        filtereditem=WOBOQItem.objects.filter(id=itemid,project=project)
        if ActiveBillExists(project):
            current_bill=Bill.objects.get(project__user=request.user,submitted=False)
        else:
            item=None
            messages.info(request,'no active bill found')
            context={            
            'item':item,
            'page_object':page_object,
            'current_quantity':current_quantity

            }
            return render(request,'project/measurements.html',context)
        if filtereditem:
            item=filtereditem[0]
            if request.method=='POST': 
                try: 
                    mdesc=request.POST.get('mdesc')          
                    count = request.POST.get('newcount')
                    count = int(count) if count else 1
                    number = request.POST.get('newnumber')
                    number = Decimal(number) if number else 1.000
                    length = request.POST.get('newlength')
                    length = Decimal(length) if length else 1.000
                    breadth = request.POST.get('newbreadth')
                    breadth = Decimal(breadth) if breadth else 1.000
                    height = request.POST.get('newheight')
                    height = Decimal(height) if height else 1.000
                    entry_date=request.POST.get('entry_date')
                    quantity=count*number*length*breadth*height
                    print("<<<<quantity=",quantity)
                    new_rec=Measurement(
                    woboqitem=item,
                    bill=current_bill,
                    count=count,
                    number=number,
                    length=length,
                    breadth=breadth,
                    height=height,
                    quantity=quantity,
                    entry_date=entry_date,
                    m_description=mdesc,
                    )
                    new_measure=new_rec.save()
                    
                    # print('Measurement added successfully.',new_measure)
                    messages.info(request,'Measurement added successfully.')
                except Exception as e:
                    messages.error(request, "Error : "+str(e))
                    # print('Error occcured during measurement update',e)
                    
            measurements=Measurement.objects.filter(woboqitem=item)
            
            
            if measurements:
                total=measurements.aggregate(qty=Sum('quantity'))
                if total['qty'] is not None:                     
                    current_quantity=total['qty'] 
            current_measurements=Measurement.objects.filter(woboqitem=item,bill=current_bill).order_by('-entry_date')
            last_measure=current_measurements.first()
            if last_measure:
                last_measure_date=last_measure.entry_date
            paginator=Paginator(current_measurements,100)
            page_object=paginator.get_page(page)
        else:
            item=None
            messages.info(request,'Item not found')
            last_measure_date=None
        
        context={
            
            'item':item,
            'page_object':page_object,
            'current_quantity':current_quantity,
            'last_measure_date':last_measure_date

        }
        return render(request,'project/measurements.html',context)
    except Exception as e:
        messages.error(request, "Error : "+str(e))
    return render(request,'project/measurements.html') 


@project_required
def edit_measurement(request,measurementid):
    msg='' 
    try:
        objs=Measurement.objects.filter(id=measurementid)
        if objs:
            measure=objs[0]
            item=measure.woboqitem
            if measure.bill.submitted:
                bill=measure.bill.pk
            else:
                bill=None
            if request.method == 'POST':
                mdesc=request.POST.get('mdesc')          
                count = request.POST.get('newcount')
                count = int(count) if count else 1
                number = request.POST.get('newnumber')
                number = Decimal(number) if number else 1.000
                length = request.POST.get('newlength')
                length = Decimal(length) if length else 1.000
                breadth = request.POST.get('newbreadth')
                breadth = Decimal(breadth) if breadth else 1.000
                height = request.POST.get('newheight')
                height = Decimal(height) if height else 1.000
                entry_date=request.POST.get('entry_date')
                new_quantity=count*number*length*breadth*height               
                measurements=Measurement.objects.filter(woboqitem=item).exclude(id=measurementid)   
                total_quantity=Decimal(0.00)   
                if measurements:
                    total=measurements.aggregate(qty=Sum('quantity'))
                    if total['qty'] is not None:                     
                        total_quantity=total['qty'] 
                available_qty=item.quantity-total_quantity
                if total_quantity+new_quantity > item.quantity:
                    print('Quantity exceeds the BoQ item quantity',item.quantity)
                    msg= f'Quantity exceeds the BoQ item quantity {item.quantity}, Available quantity is only {available_qty}{item.unit}'  
                else:
                    measure.count=count
                    measure.number=number
                    measure.length=length
                    measure.breadth=breadth
                    measure.height=height
                    measure.quantity=new_quantity
                    measure.entry_date=entry_date
                    measure.m_description=mdesc
                    measure.save()
                    measure=Measurement.objects.get(id=measurementid)
                    print('Measurement edited successfully')
                    msg='Measurement updated successfully'
                    
            context={
                'measurement':measure,
                'item':item,
                'msg':msg,
                'bill':bill
            }

        else:
            print("no measure found.")
            context={
                
                'msg':'Measurement not found,'
            }
    except Exception as e:
        messages.error(request, "Error : "+str(e))
        
        print("exception during measurement edit,",e)    
    return render(request,'project/edit_measurement.html',context)

@project_required
def remove_measurement(request,measurementid):
    try:
        objs=Measurement.objects.filter(id=measurementid)
        if objs:
            measure=objs[0]
            itemid=measure.woboqitem.pk
            measure.delete()
            # print('Measurement removed successfully')
            return redirect('project:Measurements',itemid,1)
        else:
            print("no measure found.")
    except Exception as e:
        messages.error(request, "Error : "+str(e))        
    return redirect('project:CurrentBill',1)

#managing previous bills
@project_required
def previous_bills(request):
    try:
        if not Bill.objects.filter(project__user=request.user).exists():
            messages.error(request, "No bills found.")
            return render(request,'project/previous_bills.html')
        previous_bills=Bill.objects.filter(project__user=request.user,submitted=True)
        return render(request,'project/previous_bills.html',{'bills':previous_bills})
    except Exception as e:
        messages.error(request, "Error : "+str(e))
    return render(request,'project/previous_bills.html',{'bills':None}) 



@project_required
def previous_bill(request,billid,page):  
    try:  
        msg=''
        bill=None
        currentuser=request.user
        projectobjs=Project.objects.filter(user=currentuser)
        if projectobjs:
            project=projectobjs[0]
        else:
            messages.error(request, " No project found. Add project and try again.")
            return render(request,'project/previous_bill.html')
        prevbill_objs=Bill.objects.filter(project=project,submitted=True)
        bills=Bill.objects.filter(id=billid)  
        if not prevbill_objs:
            messages.error(request, "No bills found.")
            return render(request,'project/previous_bill.html')
        if not bills:
            messages.error(request, "No bill found.")
            return render(request,'project/previous_bill.html')
        page_obj=None
        bill=bills[0]
        if bill not in prevbill_objs:
            messages.error(request, "No bills found.")
            return render(request,'project/previous_bill.html')

        if not WOBOQItem.objects.filter(project__user=currentuser).exists():                
            messages.error(request, 'No items found in BoQ, Add items to BoQ' )       
            return render(request,'project/previous_bill.html')
        allrecs=WOBOQItem.objects.filter(project=project).order_by('id')
        allrecs_paginator=Paginator(allrecs,100) 
        page_obj=allrecs_paginator.get_page(page)
        context={
            
            'bill':bill,        
            'page_object':page_obj,       
        }
        
        return render(request,'project/previous_bill.html',context)
    except Exception as e:
        messages.error(request, "Error : "+str(e))
    return render(request,'project/previous_bill.html',{'bill':None,  'page_object':None })


@project_required
def previous_bill_measurements(request,billid,itemid,page):
    try:
        msg='' 
        current_quantity = 0.00   
        page_object=None
        last_measure_date=None
        project=Project.objects.get(user=request.user)
        filtereditem=WOBOQItem.objects.filter(id=itemid,project=project)
        prevbills=Bill.objects.filter(project__user=request.user,submitted=True)
        billobjs=Bill.objects.filter(id=billid)
        if prevbills and billobjs :
            thisbill=billobjs[0]
            if thisbill not in prevbills:
                return render(request, 'project/previous_measurements.html',{'msg':'No bill found.'})      
            
        else:
            item=None
            messages.info(request,'no active bill found')
            context={
            
            'item':item,
            'page_object':page_object,
            'current_quantity':current_quantity

            }
            return render(request,'project/previous_measurements.html',context)
        if filtereditem:
            item=filtereditem[0]
            if request.method=='POST': 
                try: 
                    mdesc=request.POST.get('mdesc')          
                    count = request.POST.get('newcount')
                    count = int(count) if count else 1
                    number = request.POST.get('newnumber')
                    number = Decimal(number) if number else 1.000
                    length = request.POST.get('newlength')
                    length = Decimal(length) if length else 1.000
                    breadth = request.POST.get('newbreadth')
                    breadth = Decimal(breadth) if breadth else 1.000
                    height = request.POST.get('newheight')
                    height = Decimal(height) if height else 1.000
                    entry_date=request.POST.get('entry_date')
                    quantity=count*number*length*breadth*height
                    # print("<<<<quantity=",quantity)
                    new_rec=Measurement(
                    woboqitem=item,
                    bill=thisbill,
                    count=count,
                    number=number,
                    length=length,
                    breadth=breadth,
                    height=height,
                    quantity=quantity,
                    entry_date=entry_date,
                    m_description=mdesc,
                    )
                    new_measure=new_rec.save()
                    
                    # print('Measurement added successfully.',new_measure)
                    messages.info(request,'Measurement added successfully.')
                except Exception as e:
                    messages.error(request, "Error : "+str(e))
                    # print('Error occcured during measurement update',e)                   
                    
                    return render(request,'project/error_page.html',{'error_message': str(e)})   
            measurements=Measurement.objects.filter(woboqitem=item)
            
            
            if measurements:
                total=measurements.aggregate(qty=Sum('quantity'))
                if total['qty'] is not None:                     
                    current_quantity=total['qty'] 
            current_measurements=measurements.filter(bill=thisbill).order_by('-entry_date')
            last_measure=current_measurements.first()
            if last_measure:
                last_measure_date=last_measure.entry_date
            paginator=Paginator(current_measurements,100)
            page_object=paginator.get_page(page)
        else:
            item=None
            messages.info(request,'Item not in your boq.')
            last_measure_date=None
        
        context={
            'msg':msg,
            'item':item,
            'page_object':page_object,
            'current_quantity':current_quantity,
            'last_measure_date':last_measure_date,
            'bill':thisbill

        }
        return render(request,'project/previous_measurements.html',context)
    except Exception as e:
        messages.error(request, "Error : "+str(e))
    return render(request,'project/previous_measurements.html')  


@project_required
def remove_prev_measurement(request,billid,measurementid):
    try:
        objs=Measurement.objects.filter(id=measurementid)
        if objs:
            measure=objs[0]
            itemid=measure.woboqitem.pk
            billid=measure.bill.pk
            measure.delete()
            messages.success(request,'Measurement removed successfully')
            return redirect('project:PreviousMeasurements',billid,itemid,1)
        else:
            print("no measure found.")
    except Exception as e:
        messages.error(request, "Error : "+str(e))         

    return redirect('project:PreviousBill',billid,1)

@project_required
def search_bill_item(request,billid,page):    
    try:
        msg=''
        current_bill=None
        currentuser=request.user
        projectobjs=Project.objects.filter(user=currentuser)
        if projectobjs:
            project=projectobjs[0]
        else:
            messages.info(request,'No project found. Add project and try again.')
        if not Bill.objects.filter(project=project,id=billid).exists():
            messages.info(request,'Bill not found.')
            return render(request,'project/current_bill.html')
        bill_objs=Bill.objects.filter(project=project,id=billid)
        bill=bill_objs[0]    
        
        page_obj=None
        
        if WOBOQItem.objects.filter(project__user=currentuser).exists():
            searchterm=request.POST.get('searchitem')
            allrecs=WOBOQItem.objects.filter(project=project,description__icontains=searchterm).order_by('id')
            if allrecs:
                allrecs_paginator=Paginator(allrecs,len(allrecs)) 
                page_obj=allrecs_paginator.get_page(page)
                    
        else:        
            messages.info(request,'No items found in BoQ, Add items to BoQ')
        context={
            'msg' : msg,
            'bill':bill,        
            'page_object':page_obj,       
        }
        
        return render(request,'project/previous_bill.html',context)
    except Exception as e:
        messages.error(request, "Error : "+str(e))
    return redirect('project:PreviousBill',billid,1)
  



@project_required
def viewAbstract(request,billid,page):
    try:
        if not Bill.objects.filter(project__user=request.user,id=billid).exists():
            HttpResponseBadRequest("<h1>Bad Request:::Permission Denied.</h1>")
        billobjs=Bill.objects.filter(project__user=request.user,id=billid)
        bill=billobjs[0]
        project=bill.project
        previous_bills=Bill.objects.filter(project=project,bill_number__lt=bill.bill_number,submitted=True).order_by('bill_number')
        allitems=WOBOQItem.objects.filter(project=project).order_by('id')
        abstract_rows=[]
        #insert items
        #workbook=insert_items_to_abstract(workbook,allitems,bill,previous_bills)
        allitems_paginator=Paginator(allitems,200)
        items=allitems_paginator.get_page(page)
        total_boq_amt=Decimal(0.00)
        total_previous_amt=Decimal(0.00)
        total_present_amt=Decimal(0.00)   
        total_cum_amt=Decimal(0.00)
        total_bal_amt=Decimal(0.00) 
        
        for item in items: 
            if item.heading:
                if item.serial_no == 'nan':
                    serial_no=''
                else:
                    serial_no=item.serial_no
                abstract_rows.append([serial_no,item.description ,"","","","","","","","","","","","","","","",""])       
            else:
                boq_qty=item.quantity
                rate=item.rate
                units=item.unit
                boq_amt=item.amount
                previous_qty=Decimal(0.000)
                previous_amt=Decimal(0.000)
                present_qty=Decimal(0.000)
                present_amt=Decimal(0.000)            
                if previous_bills:
                    for prebill in previous_bills:                
                        Qsum=Measurement.objects.filter(bill=prebill,woboqitem=item).aggregate(itemqty=Sum('quantity'))
                        prebill_qty=Qsum['itemqty']
                        if not prebill_qty:
                            prebill_qty=Decimal(0.000)
                        prebill_amt=rate*prebill_qty
                        previous_qty=previous_qty+prebill_qty
                        previous_amt=previous_amt+prebill_amt
                PQsum=Measurement.objects.filter(bill=bill,woboqitem=item).aggregate(itemqty=Sum('quantity'))
                present_qty=PQsum['itemqty']
                if not  present_qty:
                    present_qty=Decimal(0.00)

                present_amt=present_qty*rate
                cum_qty=previous_qty+present_qty
                cum_amt=previous_amt+present_amt
                bal_qty=boq_qty-cum_qty
                bal_amt=boq_amt-cum_amt
                total_bal_amt=total_bal_amt+bal_amt
                total_boq_amt=total_boq_amt+boq_amt
                total_cum_amt=total_cum_amt+cum_amt
                total_present_amt=total_present_amt+present_amt
                total_previous_amt=total_previous_amt+previous_amt
                if item.serial_no == 'nan':
                    serial_no=''
                else:
                    serial_no=item.serial_no
                abstract_rows.append([serial_no,item.description,f'{boq_qty:.3f}',f'{rate:.3f}',units,f'{boq_amt:.2f}',f'{previous_qty:.3f}',f'{rate:.3f}',f'{previous_amt:.2f}',f'{present_qty:.3f}',f'{rate:.3f}', f'{present_amt:.2f}', f'{cum_qty:.3f}',f'{rate:.3f}', f'{cum_amt:.2f}', f'{bal_qty:.3f}',f'{rate:.3f}', f'{bal_amt:.2f}'])       
               
        prev_grand_total=Decimal(0.00)
        for prev_bill in previous_bills:
            prev_bill_total=get_grand_total(prev_bill)        
            prev_grand_total=prev_grand_total+prev_bill_total
        grand_total=get_grand_total(bill)
        
        total_amount = WOBOQItem.objects.filter(project=bill.project).aggregate(total_sum=models.Sum('amount'))
           
        if total_amount['total_sum'] is None:
            boq_total = 0.00        
        else:
            boq_total=total_amount['total_sum']
        
        
        cumu_grand_total=prev_grand_total+grand_total
        bal_grand_total=boq_total-cumu_grand_total
        
        abstract_rows.append(["","Grand Total","","","",f'{boq_total:.2f}',"","",f'{prev_grand_total:.2f}',"","",f'{grand_total:.2f}',"","",f'{cumu_grand_total:.2f}',"","",f'{bal_grand_total:.2f}'])        
        return render(request,'project/view_abstract.html',{'abstract_rows':abstract_rows,'page_object':items,'bill':bill,'last_page':allitems_paginator.num_pages})
    except Exception as e:
        messages.error(request, "Error : "+str(e))
        print(e)
        return render(request,'project/error_page.html',{'error_message': str(e)})   


@project_required
def generate_documents(request):
    lastbill=None
    try:
        projectobjs=Project.objects.filter(user=request.user)
        if not projectobjs:
            return redirect('project:Dashboard')
        
        project=projectobjs[0]
        
        bills=Bill.objects.filter(project=project).order_by('-bill_number')
        # print("lastbill",lastbill)
        return render(request,'project/documents.html',{'bills':bills})
    except Exception as e:
        messages.error(request, "Error : "+str(e))
        # print('exception during bill submission',e)
        return render(request,'project/error_page.html',{'error_message': str(e)})   

    

@project_required
def save_current_bill(request):
    currentuser=request.user
    try:
        projectobjs=Project.objects.filter(user=currentuser)
        project=projectobjs[0]
        currentbill_objs=Bill.objects.filter(project=project,submitted=False)
        current_bill=currentbill_objs[0]
        current_bill.submitted=True
        current_bill.save()        
        messages.info(request,'Bill closed. You can download documents here. You can create next billnow.')
        return redirect('project:Documents')
    except Exception as e:
        messages.error(request, "Error : "+str(e))
    return redirect('project:CurrentBill',1)    
        
#download measurement book as excel

def insert_to_mbook_excel(workbook,bill):
    try:
        project=bill.project
        items=WOBOQItem.objects.filter(project=project).order_by('id')    
        billed_items=WOBOQItem.objects.filter(id__in=Measurement.objects.filter(bill=bill).values_list('woboqitem', flat=True).distinct()).order_by('id')
        
        if items is None or billed_items is None:
            return workbook
        decimal_style = NamedStyle(name="two_decimal")
        decimal_style.number_format = '0.00'
        decimal_style1 = NamedStyle(name="three_decimal")
        decimal_style1.number_format = '0.000'
        sheet = workbook.active    
        for item in items:
            if item.heading:
                if item.serial_no == 'nan':
                    serial_no=''
                else:
                    serial_no=item.serial_no
                row_val=[serial_no,item.description,"","","","","","","","",""]
                sheet.append(row_val)             
                for cell in sheet[sheet.max_row]:                 
                    cell.font = Font(bold=True)
                    
            else:
                if item not in billed_items:
                    if item.serial_no == 'nan':
                        serial_no=''
                    else:
                        serial_no=item.serial_no
                    row_val=[serial_no,item.description,item.unit,"-","-","-","-","-","-","-","-"]
                    sheet.append(row_val)
                    sheet[sheet.max_row][1].alignment=Alignment(wrapText=True) 
                    continue                             
                measurements=Measurement.objects.filter(woboqitem=item,bill=bill).order_by('-entry_date')
                                
                itemquantity=measurements.aggregate(totalqty=Sum('quantity'))
                totalqtystr=f"{itemquantity['totalqty']:.3f} {item.unit}"
                row_val=[item.serial_no,item.description,item.unit,"","","","","","","",""]
                sheet.append(row_val)
                sheet[sheet.max_row][1].alignment=Alignment(wrapText=True) 
                for measurement in measurements:
                    row_val=["","","",measurement.entry_date,measurement.m_description,measurement.number,measurement.count,measurement.length,measurement.breadth,measurement.height,measurement.quantity]
                    sheet.append(row_val)
                    sheet[sheet.max_row][4].alignment=Alignment(wrapText=True) 
                    sheet.cell(row=sheet.max_row, column=8).style = decimal_style1 
                    sheet.cell(row=sheet.max_row, column=9).style = decimal_style1 
                    sheet.cell(row=sheet.max_row, column=10).style = decimal_style1 
                    sheet.cell(row=sheet.max_row, column=11).style = decimal_style1 
                    row_val=[]
                row_val=["","","","","","","","","","Total Quantity :",totalqtystr]
                sheet.append(row_val)
                
                for cell in sheet[sheet.max_row]:                 
                    cell.font = Font(bold=True) 
                    cell.alignment=Alignment(horizontal='right')  
                            
                measurements=[] 
    except Exception as e:
        pass
           
    return workbook

@project_required
def downloadMBook(request,billid):
    try:
        if not Bill.objects.filter(project__user=request.user,id=billid).exists():
            HttpResponseBadRequest("<h1>Bad Request:::Permission Denied.</h1>")
        bill=Bill.objects.filter(project__user=request.user,id=billid)
        workbook = openpyxl.Workbook()
        bold_font = Font(bold=True)   
        
        fill=PatternFill(fill_type='solid', fgColor='e6e6e6')  # background color
        
        sheet = workbook.active
        sheet.column_dimensions['B'].width = 75         
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['J'].width = 20
        sheet.column_dimensions['K'].width = 20
        a1=sheet.cell(row=1, column=1, value='Serial No')
        b1=sheet.cell(row=1, column=2, value='Item Description')
        c1=sheet.cell(row=1, column=3, value='Unit')
        d1=sheet.cell(row=1, column=4, value='Date')
        e1=sheet.cell(row=1, column=5, value='Description')
        f1=sheet.cell(row=1, column=6, value='No.s')
        g1=sheet.cell(row=1, column=7, value='Count')
        h1=sheet.cell(row=1, column=8, value='Length')
        i1=sheet.cell(row=1, column=9, value='Breadth')
        j1=sheet.cell(row=1, column=10, value='Height')
        k1=sheet.cell(row=1, column=11, value='Quantity')
        a1.font=b1.font=c1.font=d1.font=e1.font=f1.font=g1.font=h1.font=bold_font
        i1.font=j1.font=k1.font=bold_font
        bill=bill[0]
        
        workbook=insert_to_mbook_excel(workbook,bill)
        response=HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # Set the file name
        response['Content-Disposition'] = 'attachment; filename="measurement_book.xlsx"'
        workbook.save(response)
        return response
    except Exception as e:
        messages.error(request, "Error : "+str(e))
        return render(request,'project/error_page.html',{'error_message': str(e)})     
    


def insert_items_to_abstract(workbook,items,bill,previous_bills):
    decimal_style = NamedStyle(name="two_decimal")
    decimal_style.number_format = '0.00'
    decimal_style1 = NamedStyle(name="three_decimal")
    decimal_style1.number_format = '0.000'
    if items is None:
        return workbook 
    sheet = workbook.active
    total_boq_amt=Decimal(0.00)
    total_previous_amt=Decimal(0.00)
    total_present_amt=Decimal(0.00)   
    total_cum_amt=Decimal(0.00)
    total_bal_amt=Decimal(0.00) 
    for item in items: 
        if item.heading:
            if item.serial_no == 'nan':
                serial_no=''
            else:
                serial_no=item.serial_no
            row_val=[serial_no,"","","","","","","","","","","","","","","","",""]        
            sheet.append(row_val)
            # last row number
            last_row = sheet.max_row
            merge_range = f"B{last_row}:R{last_row}"
            sheet.merge_cells(merge_range)
            sheet[f"B{last_row}"].value = item.description   
            sheet[f"B{last_row}"].alignment = Alignment(wrap_text=True,vertical="center")
            sheet[f"B{last_row}"].font=Font(bold=True)
            
        else:
            boq_qty=item.quantity
            rate=item.rate
            units=item.unit
            boq_amt=item.amount
            previous_qty=Decimal(0.000)
            previous_amt=Decimal(0.000)
            present_qty=Decimal(0.000)
            present_amt=Decimal(0.000)            
            if previous_bills:
                for prebill in previous_bills:                
                    Qsum=Measurement.objects.filter(bill=prebill,woboqitem=item).aggregate(itemqty=Sum('quantity'))
                    prebill_qty=Qsum['itemqty']
                    if not prebill_qty:
                        prebill_qty=Decimal(0.000)
                    prebill_amt=rate*prebill_qty
                    previous_qty=previous_qty+prebill_qty
                    previous_amt=previous_amt+prebill_amt
            PQsum=Measurement.objects.filter(bill=bill,woboqitem=item).aggregate(itemqty=Sum('quantity'))
            present_qty=PQsum['itemqty']
            if not  present_qty:
                present_qty=Decimal(0.00)

            present_amt=present_qty*rate
            cum_qty=previous_qty+present_qty
            cum_amt=previous_amt+present_amt
            bal_qty=boq_qty-cum_qty
            bal_amt=boq_amt-cum_amt
            total_bal_amt=total_bal_amt+bal_amt
            total_boq_amt=total_boq_amt+boq_amt
            total_cum_amt=total_cum_amt+cum_amt
            total_present_amt=total_present_amt+present_amt
            total_previous_amt=total_previous_amt+previous_amt
            if item.serial_no == 'nan':
                serial_no=''
            else:
                serial_no=item.serial_no
            row_val=[serial_no,item.description,boq_qty.quantize(Decimal('0.000')),rate.quantize(Decimal('0.000')),units,boq_amt.quantize(Decimal('0.00')),previous_qty.quantize(Decimal('0.000')),rate.quantize(Decimal('0.000')),previous_amt.quantize(Decimal('0.00')),present_qty.quantize(Decimal('0.000')),rate.quantize(Decimal('0.000')),present_amt.quantize(Decimal('0.00')),cum_qty.quantize(Decimal('0.000')),rate.quantize(Decimal('0.000')),cum_amt.quantize(Decimal('0.00')),bal_qty.quantize(Decimal('0.000')),rate.quantize(Decimal('0.000')),bal_amt.quantize(Decimal('0.00'))]        
            sheet.append(row_val)
            sheet.cell(row=sheet.max_row, column=3).style = decimal_style1 
            sheet.cell(row=sheet.max_row, column=4).style = decimal_style1 
            sheet.cell(row=sheet.max_row, column=6).style = decimal_style 

            sheet.cell(row=sheet.max_row, column=7).style = decimal_style1 
            sheet.cell(row=sheet.max_row, column=8).style = decimal_style1 
            sheet.cell(row=sheet.max_row, column=9).style = decimal_style 

            sheet.cell(row=sheet.max_row, column=10).style = decimal_style1 
            sheet.cell(row=sheet.max_row, column=11).style = decimal_style1 
            sheet.cell(row=sheet.max_row, column=12).style = decimal_style

            sheet.cell(row=sheet.max_row, column=13).style = decimal_style1 
            sheet.cell(row=sheet.max_row, column=14).style = decimal_style1 
            sheet.cell(row=sheet.max_row, column=15).style = decimal_style  

            sheet.cell(row=sheet.max_row, column=16).style = decimal_style1 
            sheet.cell(row=sheet.max_row, column=17).style = decimal_style1 
            sheet.cell(row=sheet.max_row, column=18).style = decimal_style  
                
    return workbook



def insert_subheading(workbook,subheading):
    sheet = workbook.active
    parts=subheading.split('&&&')
    row_val=[parts[0],parts[1],"","","","","","","","","","","","","","","",""]        
    sheet.append(row_val)
    for cell in sheet[sheet.max_row]:                 
        cell.font = Font(bold=True)
        cell.fill=PatternFill(fill_type='solid', fgColor='e6e6e6')
    return workbook
    
